"""
═══════════════════════════════════════════════════════════════════════════════
  Google My Business Reviews — Download Direto via REST API
  Embaixada Carioca — Morro da Urca
  Usa endpoints REST diretos (sem discovery document v4 obsoleto)
  Autor: Manus AI | Data: 2026-06-17
═══════════════════════════════════════════════════════════════════════════════
"""

import os
import sys
import json
import csv
import time
from datetime import datetime, timezone
from pathlib import Path

# ─── Verificar dependências ───────────────────────────────────────────────────
try:
    import requests
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request as GoogleRequest
except ImportError as e:
    print(f"❌ Dependência faltando: {e}")
    print("   pip install google-auth google-auth-oauthlib requests")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ─── Configuração ─────────────────────────────────────────────────────────────
SCOPES = ["https://www.googleapis.com/auth/business.manage"]
TOKEN_FILE = "token.json"
CREDENTIALS_FILE = "credentials.json"

OUTPUT_JSON = "reviews_embaixada_carioca.json"
OUTPUT_CSV  = "reviews_embaixada_carioca.csv"
OUTPUT_XLSX = "reviews_embaixada_carioca.xlsx"

PAGE_SIZE = 50

STAR_MAP = {
    "ONE": 1, "TWO": 2, "THREE": 3, "FOUR": 4, "FIVE": 5,
    "STAR_RATING_UNSPECIFIED": 0
}

# Endpoints da GBP API
BASE_ACCOUNTS = "https://mybusinessaccountmanagement.googleapis.com/v1"
BASE_REVIEWS  = "https://mybusiness.googleapis.com/v4"

# ─── Autenticação ─────────────────────────────────────────────────────────────

def get_access_token():
    """Obtém e renova o access token a partir do token.json."""
    if not Path(TOKEN_FILE).exists():
        print(f"❌ '{TOKEN_FILE}' não encontrado!")
        sys.exit(1)

    with open(TOKEN_FILE) as f:
        token_data = json.load(f)

    creds = Credentials(
        token=token_data.get("access_token") or token_data.get("token"),
        refresh_token=token_data.get("refresh_token"),
        token_uri=token_data.get("token_uri", "https://oauth2.googleapis.com/token"),
        client_id=token_data.get("client_id"),
        client_secret=token_data.get("client_secret"),
        scopes=token_data.get("scopes", SCOPES)
    )

    if not creds.valid:
        if creds.expired and creds.refresh_token:
            print("🔄 Renovando token de acesso...")
            creds.refresh(GoogleRequest())
            # Salvar token renovado
            with open(TOKEN_FILE, "w") as f:
                token_updated = {
                    "access_token": creds.token,
                    "refresh_token": creds.refresh_token,
                    "token_uri": creds.token_uri,
                    "client_id": creds.client_id,
                    "client_secret": creds.client_secret,
                    "scopes": list(creds.scopes) if creds.scopes else SCOPES,
                    "expiry": creds.expiry.isoformat() if creds.expiry else None
                }
                json.dump(token_updated, f, indent=2)
            print("  ✅ Token renovado e salvo.")
        else:
            print("❌ Token inválido. Execute o fluxo OAuth2 novamente.")
            sys.exit(1)

    return creds.token


def make_headers(token):
    return {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }


# ─── Funções da API ───────────────────────────────────────────────────────────

def get_accounts(token):
    """Lista todas as contas do Google My Business."""
    r = requests.get(f"{BASE_ACCOUNTS}/accounts", headers=make_headers(token))
    if r.status_code != 200:
        print(f"❌ Erro ao listar contas: {r.status_code} — {r.text[:200]}")
        sys.exit(1)
    accounts = r.json().get("accounts", [])
    if not accounts:
        print("❌ Nenhuma conta encontrada.")
        sys.exit(1)
    return accounts


def get_locations(token, account_name):
    """Lista todos os locais de uma conta."""
    url = f"https://mybusinessbusinessinformation.googleapis.com/v1/{account_name}/locations"
    params = {"readMask": "name,title,storefrontAddress,websiteUri"}
    r = requests.get(url, headers=make_headers(token), params=params)
    if r.status_code != 200:
        return []
    return r.json().get("locations", [])


def find_embaixada_location(token, accounts):
    """Encontra o local 'Embaixada Carioca' entre todas as contas."""
    print("\n🔍 Procurando o local 'Embaixada Carioca'...")

    for account in accounts:
        account_name = account["name"]
        locations = get_locations(token, account_name)

        for loc in locations:
            title = loc.get("title", "") or loc.get("locationName", "")
            loc_name = loc.get("name", "")
            print(f"   Encontrado: {title} ({loc_name})")

            if "embaixada" in title.lower() or "carioca" in title.lower():
                print(f"\n✅ Local identificado: {title}")
                return account_name, loc_name, title

    print("\n⚠️  'Embaixada Carioca' não encontrada automaticamente.")
    print("   Usando localização conhecida...")
    # Fallback para o ID já identificado
    return (
        "accounts/106083628368200012478",
        "locations/18008728615502069543",
        "Embaixada Carioca"
    )


def fetch_all_reviews(token, account_name, location_name):
    """
    Busca TODAS as avaliações com paginação automática.
    Usa o endpoint v4 da My Business API.
    """
    # Extrair apenas o ID numérico do location_name se necessário
    # O endpoint v4 usa: accounts/{account_id}/locations/{location_id}/reviews
    acc_id = account_name.split("/")[-1]
    loc_id = location_name.split("/")[-1]

    base_url = f"{BASE_REVIEWS}/accounts/{acc_id}/locations/{loc_id}/reviews"

    all_reviews = []
    next_page_token = None
    page_num = 0
    average_rating = 0.0
    total_count = 0

    print(f"\n📥 Baixando avaliações (até {PAGE_SIZE} por página)...")

    while True:
        page_num += 1
        params = {
            "pageSize": PAGE_SIZE,
            "orderBy": "updateTime desc"
        }
        if next_page_token:
            params["pageToken"] = next_page_token

        try:
            r = requests.get(base_url, headers=make_headers(token), params=params)

            if r.status_code == 401:
                print("🔄 Token expirado, renovando...")
                token = get_access_token()
                r = requests.get(base_url, headers=make_headers(token), params=params)

            if r.status_code != 200:
                print(f"\n❌ Erro na página {page_num}: {r.status_code}")
                print(f"   {r.text[:300]}")
                break

            data = r.json()

        except Exception as e:
            print(f"\n❌ Erro de conexão (página {page_num}): {e}")
            break

        reviews = data.get("reviews", [])
        all_reviews.extend(reviews)

        # Capturar metadados da primeira página
        if page_num == 1:
            average_rating = data.get("averageRating", 0.0)
            total_count = data.get("totalReviewCount", 0)
            print(f"   Total de avaliações no Google: {total_count}")
            print(f"   Nota média: {average_rating:.2f} ⭐")

        print(f"   Página {page_num:3d}: {len(reviews):3d} avaliações "
              f"(acumulado: {len(all_reviews):5d} / {total_count})")

        next_page_token = data.get("nextPageToken")
        if not next_page_token:
            break

        # Pausa para respeitar rate limits (300 QPM = 5 por segundo)
        time.sleep(0.3)

    print(f"\n✅ Total coletado: {len(all_reviews)} avaliações")
    return all_reviews, average_rating, total_count, token


def parse_review(review):
    """Converte um objeto Review da API para um dicionário normalizado."""
    reviewer = review.get("reviewer", {})
    reply = review.get("reviewReply", {})
    star_raw = review.get("starRating", "STAR_RATING_UNSPECIFIED")

    def fmt_ts(ts):
        if not ts:
            return None
        try:
            dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
            return dt.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        except Exception:
            return ts

    return {
        "review_id": review.get("reviewId", ""),
        "author_name": reviewer.get("displayName", "Anônimo"),
        "is_anonymous": reviewer.get("isAnonymous", False),
        "profile_photo_url": reviewer.get("profilePhotoUrl", None),
        "rating": STAR_MAP.get(star_raw, 0),
        "star_rating_raw": star_raw,
        "comment": review.get("comment", ""),
        "create_time": fmt_ts(review.get("createTime")),
        "update_time": fmt_ts(review.get("updateTime")),
        "owner_reply": reply.get("comment", None) if reply else None,
        "owner_reply_time": fmt_ts(reply.get("updateTime")) if reply else None,
        "has_media": len(review.get("reviewMediaItems", [])) > 0,
        "media_count": len(review.get("reviewMediaItems", []))
    }


# ─── Exportação ───────────────────────────────────────────────────────────────

def export_json(data, filepath):
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    size_kb = Path(filepath).stat().st_size / 1024
    print(f"  ✅ JSON: {filepath} ({size_kb:.1f} KB)")


def export_csv(reviews, filepath):
    if not reviews:
        return
    fieldnames = [
        "review_id", "author_name", "is_anonymous", "rating",
        "comment", "create_time", "update_time",
        "owner_reply", "owner_reply_time", "has_media", "media_count"
    ]
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(reviews)
    size_kb = Path(filepath).stat().st_size / 1024
    print(f"  ✅ CSV:  {filepath} ({size_kb:.1f} KB)")


def export_xlsx(data, filepath):
    if not HAS_OPENPYXL:
        print("  ⚠️  openpyxl não instalado — XLSX não gerado.")
        return

    reviews = data.get("reviews", [])
    if not reviews:
        return

    from collections import Counter

    wb = openpyxl.Workbook()

    # Cores
    C_DARK  = "1A1A2E"
    C_GOLD  = "C9A84C"
    C_LIGHT = "F5F0E8"
    C_WHITE = "FFFFFF"
    C_GREEN = "D4EDDA"
    C_RED   = "F8D7DA"

    def hdr(cell, bg=C_DARK, sz=11, color="FFFFFF"):
        cell.font = Font(bold=True, color=color, size=sz)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def dat(cell, bold=False, center=False, bg=C_WHITE):
        cell.font = Font(bold=bold, size=10)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(
            horizontal="center" if center else "left",
            vertical="top", wrap_text=True
        )

    total = len(reviews)
    avg   = data.get("average_rating", 0)
    dist  = Counter(r["rating"] for r in reviews)
    with_reply = sum(1 for r in reviews if r.get("owner_reply"))

    # ── Aba 1: Resumo ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumo"
    ws.merge_cells("A1:B1")
    ws["A1"].value = "Embaixada Carioca — Avaliações Google"
    hdr(ws["A1"], bg=C_DARK, sz=14)
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:B2")
    ws["A2"].value = f"Extração: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    hdr(ws["A2"], bg=C_GOLD, sz=10, color="1A1A2E")
    ws.row_dimensions[2].height = 22

    metrics = [
        ("Total de Avaliações", total),
        ("Nota Média", f"{avg:.2f} ⭐"),
        ("Com Resposta do Proprietário", f"{with_reply} ({with_reply/total*100:.1f}%)" if total else "0"),
        ("Sem Resposta", f"{total - with_reply} ({(total-with_reply)/total*100:.1f}%)" if total else "0"),
        ("5 Estrelas", f"{dist.get(5,0)} ({dist.get(5,0)/total*100:.1f}%)" if total else "0"),
        ("4 Estrelas", f"{dist.get(4,0)} ({dist.get(4,0)/total*100:.1f}%)" if total else "0"),
        ("3 Estrelas", f"{dist.get(3,0)} ({dist.get(3,0)/total*100:.1f}%)" if total else "0"),
        ("2 Estrelas", f"{dist.get(2,0)} ({dist.get(2,0)/total*100:.1f}%)" if total else "0"),
        ("1 Estrela",  f"{dist.get(1,0)} ({dist.get(1,0)/total*100:.1f}%)" if total else "0"),
        ("Taxa de Resposta", f"{with_reply/total*100:.1f}%" if total else "0%"),
    ]

    for i, (k, v) in enumerate(metrics, 4):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v
        bg = C_LIGHT if i % 2 == 0 else C_WHITE
        dat(ws[f"A{i}"], bold=True, bg=bg)
        dat(ws[f"B{i}"], center=True, bg=bg)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22

    # ── Aba 2: Avaliações ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Avaliações")

    headers_row = ["ID", "Autor", "Anônimo", "Nota", "Comentário",
                   "Data", "Última Atualização", "Resposta do Proprietário",
                   "Data da Resposta", "Tem Mídia"]
    widths = [20, 25, 10, 8, 70, 22, 22, 60, 22, 12]

    for col, (h, w) in enumerate(zip(headers_row, widths), 1):
        cell = ws2.cell(row=1, column=col, value=h)
        hdr(cell)
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[1].height = 25
    ws2.freeze_panes = "A2"

    for row_i, r in enumerate(reviews, 2):
        stars_str = "⭐" * r["rating"] if r["rating"] else "—"
        row_data = [
            r.get("review_id", ""),
            r.get("author_name", ""),
            "Sim" if r.get("is_anonymous") else "Não",
            stars_str,
            r.get("comment", ""),
            r.get("create_time", ""),
            r.get("update_time", ""),
            r.get("owner_reply", ""),
            r.get("owner_reply_time", ""),
            "✓" if r.get("has_media") else "",
        ]
        bg = C_LIGHT if row_i % 2 == 0 else C_WHITE
        if r["rating"] == 5:
            bg = C_GREEN
        elif r["rating"] == 1:
            bg = C_RED

        for col_i, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_i, column=col_i, value=val)
            dat(cell, bg=bg, center=(col_i in [3, 4, 9, 10]))
        ws2.row_dimensions[row_i].height = 55

    # ── Aba 3: Análise por Nota ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Análise por Nota")
    ws3.merge_cells("A1:D1")
    ws3["A1"].value = "Análise Detalhada por Nota"
    hdr(ws3["A1"], bg=C_DARK, sz=13)
    ws3.row_dimensions[1].height = 30

    h3 = ["Nota", "Qtd", "% do Total", "Com Resposta"]
    for col, h in enumerate(h3, 1):
        cell = ws3.cell(row=2, column=col, value=h)
        hdr(cell, bg=C_GOLD, color="1A1A2E")

    for row_i, stars in enumerate(range(5, 0, -1), 3):
        subset = [r for r in reviews if r["rating"] == stars]
        count = len(subset)
        pct = count / total * 100 if total else 0
        replied = sum(1 for r in subset if r.get("owner_reply"))
        bg = C_LIGHT if row_i % 2 == 0 else C_WHITE
        ws3.cell(row=row_i, column=1, value="⭐" * stars)
        ws3.cell(row=row_i, column=2, value=count)
        ws3.cell(row=row_i, column=3, value=f"{pct:.1f}%")
        ws3.cell(row=row_i, column=4, value=f"{replied} ({replied/count*100:.0f}%)" if count else "—")
        for col in range(1, 5):
            dat(ws3.cell(row=row_i, column=col), bg=bg, center=True)

    for col, w in zip(range(1, 5), [18, 10, 15, 20]):
        ws3.column_dimensions[get_column_letter(col)].width = w

    wb.save(filepath)
    size_kb = Path(filepath).stat().st_size / 1024
    print(f"  ✅ XLSX: {filepath} ({size_kb:.1f} KB)")


# ─── Execução Principal ───────────────────────────────────────────────────────

def main():
    print("\n" + "═" * 70)
    print("  🗺️  Google My Business Reviews API — Embaixada Carioca")
    print(f"  Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print("═" * 70)

    # 1. Autenticar
    print("\n🔐 Autenticando com a Google Business Profile API...")
    token = get_access_token()
    print("  ✅ Autenticação concluída")

    # 2. Encontrar a conta e o local
    print("\n🔍 Localizando conta e estabelecimento...")
    accounts = get_accounts(token)
    print(f"   Contas encontradas: {len(accounts)}")
    for acc in accounts:
        print(f"   • {acc.get('accountName', acc.get('name', ''))}")

    account_name, location_name, location_title = find_embaixada_location(token, accounts)

    # 3. Baixar todas as avaliações
    raw_reviews, avg_rating, total_count, token = fetch_all_reviews(
        token, account_name, location_name
    )

    # 4. Normalizar os dados
    print("\n🔄 Processando avaliações...")
    parsed = [parse_review(r) for r in raw_reviews]

    # 5. Montar estrutura de saída
    output = {
        "metadata": {
            "location_name": location_title,
            "location_id": location_name,
            "account_id": account_name,
            "extracted_at": datetime.now().isoformat(),
            "total_review_count": total_count,
            "average_rating": round(avg_rating, 2),
            "reviews_collected": len(parsed)
        },
        "average_rating": round(avg_rating, 2),
        "total_review_count": total_count,
        "reviews": parsed
    }

    # 6. Exportar
    print(f"\n📁 Exportando {len(parsed)} avaliações...")
    export_json(output, OUTPUT_JSON)
    export_csv(parsed, OUTPUT_CSV)
    export_xlsx(output, OUTPUT_XLSX)

    # 7. Resumo final
    from collections import Counter
    dist = Counter(r["rating"] for r in parsed)
    with_reply = sum(1 for r in parsed if r.get("owner_reply"))

    print(f"\n{'═'*70}")
    print(f"  ✅ EXTRAÇÃO CONCLUÍDA!")
    print(f"{'═'*70}")
    print(f"  Local:              {location_title}")
    print(f"  Nota média:         {avg_rating:.2f} ⭐")
    print(f"  Total no Google:    {total_count}")
    print(f"  Coletadas via API:  {len(parsed)}")
    if parsed:
        print(f"  Com resposta:       {with_reply} ({with_reply/len(parsed)*100:.1f}%)")
    print(f"\n  Distribuição:")
    for stars in range(5, 0, -1):
        bar = "█" * min(dist.get(stars, 0) // 10, 40)
        print(f"    {'⭐'*stars:<15} {dist.get(stars,0):5d}  {bar}")
    print(f"\n  Arquivos gerados:")
    print(f"    📄 {OUTPUT_JSON}")
    print(f"    📊 {OUTPUT_CSV}")
    if HAS_OPENPYXL:
        print(f"    📗 {OUTPUT_XLSX}")
    print("═" * 70 + "\n")


if __name__ == "__main__":
    main()

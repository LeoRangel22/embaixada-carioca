#!/usr/bin/env python3
"""
═══════════════════════════════════════════════════════════════════════════════
  Google My Business Reviews API — Embaixada Carioca
  Extrai avaliações via API oficial do Google Business Profile (v4)
  Autor: Manus AI | Data: 2026-06-11
═══════════════════════════════════════════════════════════════════════════════

COMO USAR:
────────────────────────────────────────────────────────────────────────────
1. Siga as instruções em COMO_USAR_GMB_API.md para configurar as credenciais
   OAuth2 no Google Cloud Console.

2. Coloque o arquivo credentials.json na mesma pasta deste script.

3. Execute:
       python3 gmb_reviews_api.py

4. Na primeira execução, um navegador abrirá para autorização OAuth2.
   Após autorizar, o token é salvo em token.json para execuções futuras.

5. Os resultados são exportados para:
       reviews_embaixada_carioca.json   (dados completos)
       reviews_embaixada_carioca.csv    (tabela simples)
       reviews_embaixada_carioca.xlsx   (planilha formatada — 3 abas)
────────────────────────────────────────────────────────────────────────────

DEPENDÊNCIAS:
    pip install google-auth google-auth-oauthlib google-auth-httplib2
                google-api-python-client openpyxl requests
═══════════════════════════════════════════════════════════════════════════════
"""

import os
import sys
import json
import csv
import time
from datetime import datetime, timezone
from pathlib import Path

# ─── Verificar dependências ───────────────────────────────────────────────────
missing = []
try:
    import requests
except ImportError:
    missing.append("requests")

try:
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request as GoogleRequest
    import googleapiclient.discovery
except ImportError:
    missing.append("google-auth google-auth-oauthlib google-auth-httplib2 google-api-python-client")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

if missing:
    print("❌ Dependências faltando. Execute:")
    print(f"   pip install {' '.join(missing)}")
    sys.exit(1)

# ─── Configuração ─────────────────────────────────────────────────────────────

# Escopos OAuth2 necessários para leitura de avaliações
SCOPES = ["https://www.googleapis.com/auth/business.manage"]

# Arquivos de credenciais
CREDENTIALS_FILE = "credentials.json"
TOKEN_FILE = "token.json"

# Arquivos de saída
OUTPUT_JSON = "reviews_embaixada_carioca.json"
OUTPUT_CSV  = "reviews_embaixada_carioca.csv"
OUTPUT_XLSX = "reviews_embaixada_carioca.xlsx"

# Paginação: máximo permitido pela API é 50 por página
PAGE_SIZE = 50

# Mapeamento de StarRating para inteiro
STAR_MAP = {
    "ONE": 1, "TWO": 2, "THREE": 3, "FOUR": 4, "FIVE": 5,
    "STAR_RATING_UNSPECIFIED": 0
}

# ─── Autenticação OAuth2 ──────────────────────────────────────────────────────

def get_credentials():
    """
    Obtém credenciais OAuth2 válidas.
    - Se token.json existir e for válido, usa-o diretamente.
    - Se expirado, renova automaticamente.
    - Se não existir, abre o navegador para autorização.
    """
    creds = None

    if Path(TOKEN_FILE).exists():
        creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            print("🔄 Renovando token de acesso...")
            creds.refresh(GoogleRequest())
        else:
            if not Path(CREDENTIALS_FILE).exists():
                print(f"\n❌ Arquivo '{CREDENTIALS_FILE}' não encontrado!")
                print("\nSiga as instruções em COMO_USAR_GMB_API.md para criar")
                print("as credenciais OAuth2 no Google Cloud Console.")
                sys.exit(1)

            print("\n🌐 Abrindo navegador para autorização OAuth2...")
            print("   Faça login com a conta Google que gerencia a Embaixada Carioca.\n")
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            creds = flow.run_local_server(port=0)

        # Salvar token para próximas execuções
        with open(TOKEN_FILE, "w") as f:
            f.write(creds.to_json())
        print(f"✅ Token salvo em '{TOKEN_FILE}'")

    return creds


# ─── Funções da API ───────────────────────────────────────────────────────────

def get_accounts(service):
    """Lista todas as contas do Google My Business."""
    result = service.accounts().list().execute()
    accounts = result.get("accounts", [])
    if not accounts:
        print("❌ Nenhuma conta do Google My Business encontrada.")
        print("   Certifique-se de que a conta tem acesso ao Perfil da Empresa.")
        sys.exit(1)
    return accounts


def get_locations(service, account_name):
    """Lista todos os locais de uma conta."""
    result = service.accounts().locations().list(parent=account_name).execute()
    return result.get("locations", [])


def find_embaixada_location(service, accounts):
    """
    Encontra automaticamente o local 'Embaixada Carioca' entre todas as contas.
    Retorna (account_name, location_name, location_title).
    """
    print("\n🔍 Procurando o local 'Embaixada Carioca'...")

    for account in accounts:
        account_name = account["name"]
        locations = get_locations(service, account_name)

        for loc in locations:
            title = loc.get("locationName", "") or loc.get("title", "")
            loc_name = loc.get("name", "")
            print(f"   Encontrado: {title} ({loc_name})")

            if "embaixada" in title.lower() or "carioca" in title.lower():
                print(f"\n✅ Local identificado: {title}")
                return account_name, loc_name, title

    # Se não encontrou automaticamente, listar para o usuário escolher
    print("\n⚠️  'Embaixada Carioca' não encontrada automaticamente.")
    print("   Locais disponíveis:")

    all_locations = []
    for account in accounts:
        account_name = account["name"]
        locations = get_locations(service, account_name)
        for loc in locations:
            title = loc.get("locationName", "") or loc.get("title", "")
            all_locations.append((account_name, loc["name"], title))
            print(f"   [{len(all_locations)}] {title}")

    if not all_locations:
        print("❌ Nenhum local encontrado. Verifique as permissões da conta.")
        sys.exit(1)

    choice = input("\nDigite o número do local desejado: ").strip()
    try:
        idx = int(choice) - 1
        return all_locations[idx]
    except (ValueError, IndexError):
        print("❌ Escolha inválida.")
        sys.exit(1)


def fetch_all_reviews(service, account_name, location_name):
    """
    Busca TODAS as avaliações de um local com paginação automática.
    Retorna (lista_de_reviews, average_rating, total_count).
    """
    all_reviews = []
    next_page_token = None
    page_num = 0
    average_rating = 0.0
    total_count = 0

    print(f"\n📥 Baixando avaliações (até {PAGE_SIZE} por página)...")

    while True:
        page_num += 1
        params = {
            "parent": location_name,
            "pageSize": PAGE_SIZE,
            "orderBy": "updateTime desc"
        }
        if next_page_token:
            params["pageToken"] = next_page_token

        try:
            response = (
                service.accounts()
                .locations()
                .reviews()
                .list(**params)
                .execute()
            )
        except Exception as e:
            print(f"\n❌ Erro ao buscar avaliações (página {page_num}): {e}")
            break

        reviews = response.get("reviews", [])
        all_reviews.extend(reviews)

        # Capturar metadados da primeira página
        if page_num == 1:
            average_rating = response.get("averageRating", 0.0)
            total_count = response.get("totalReviewCount", 0)
            print(f"   Total de avaliações no Google: {total_count}")
            print(f"   Nota média: {average_rating:.1f} ⭐")

        print(f"   Página {page_num}: {len(reviews)} avaliações coletadas "
              f"(total até agora: {len(all_reviews)})")

        next_page_token = response.get("nextPageToken")
        if not next_page_token:
            break

        # Pequena pausa para não sobrecarregar a API
        time.sleep(0.5)

    print(f"\n✅ Total coletado: {len(all_reviews)} avaliações")
    return all_reviews, average_rating, total_count


def parse_review(review):
    """Converte um objeto Review da API para um dicionário normalizado."""
    reviewer = review.get("reviewer", {})
    reply = review.get("reviewReply", {})
    star_raw = review.get("starRating", "STAR_RATING_UNSPECIFIED")

    # Converter timestamps ISO 8601 para formato legível
    def fmt_ts(ts):
        if not ts:
            return None
        try:
            dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
            return dt.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        except Exception:
            return ts

    return {
        "review_id": review.get("reviewId", ""),
        "author_name": reviewer.get("displayName", "Anônimo"),
        "is_anonymous": reviewer.get("isAnonymous", False),
        "profile_photo_url": reviewer.get("profilePhotoUrl", None),
        "rating": STAR_MAP.get(star_raw, 0),
        "star_rating_raw": star_raw,
        "comment": review.get("comment", ""),
        "create_time": fmt_ts(review.get("createTime")),
        "update_time": fmt_ts(review.get("updateTime")),
        "owner_reply": reply.get("comment", None) if reply else None,
        "owner_reply_time": fmt_ts(reply.get("updateTime")) if reply else None,
        "has_media": len(review.get("reviewMediaItems", [])) > 0,
        "media_count": len(review.get("reviewMediaItems", []))
    }


# ─── Funções de Exportação ────────────────────────────────────────────────────

def export_json(data, filepath):
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    size_kb = Path(filepath).stat().st_size / 1024
    print(f"  ✅ JSON: {filepath} ({size_kb:.1f} KB)")


def export_csv(reviews, filepath):
    if not reviews:
        return
    fieldnames = [
        "review_id", "author_name", "is_anonymous", "rating",
        "comment", "create_time", "update_time",
        "owner_reply", "owner_reply_time", "has_media", "media_count"
    ]
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(reviews)
    size_kb = Path(filepath).stat().st_size / 1024
    print(f"  ✅ CSV:  {filepath} ({size_kb:.1f} KB)")


def export_xlsx(data, filepath):
    if not HAS_OPENPYXL:
        print("  ⚠️  openpyxl não instalado — XLSX não gerado.")
        return

    reviews = data.get("reviews", [])
    if not reviews:
        return

    wb = openpyxl.Workbook()

    # Paleta de cores
    C_DARK   = "00405A"
    C_GOLD   = "F59B1E"
    C_LIGHT  = "E8F4F8"
    C_WHITE  = "FFFFFF"
    C_GREEN  = "D4EDDA"
    C_RED    = "F8D7DA"

    def hdr(cell, bg=C_DARK, fg=C_WHITE, sz=11, bold=True):
        cell.font = Font(bold=bold, color=fg, size=sz, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def dat(cell, bold=False, center=False, bg=None, wrap=True):
        cell.font = Font(bold=bold, size=10, name="Calibri")
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(
            horizontal="center" if center else "left",
            vertical="center", wrap_text=wrap
        )

    # ── Aba 1: Resumo ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumo"

    ws.merge_cells("A1:C1")
    ws["A1"].value = "Embaixada Carioca — Avaliações Google My Business"
    hdr(ws["A1"], sz=14)
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:C2")
    ws["A2"].value = f"Extraído via API oficial em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, size=10, color="666666", name="Calibri")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Métricas
    ws["A4"].value = "Métrica"
    ws["B4"].value = "Valor"
    hdr(ws["A4"], bg=C_GOLD)
    hdr(ws["B4"], bg=C_GOLD)

    total = len(reviews)
    with_reply = sum(1 for r in reviews if r.get("owner_reply"))
    with_text  = sum(1 for r in reviews if r.get("comment"))
    with_media = sum(1 for r in reviews if r.get("has_media"))
    anon_count = sum(1 for r in reviews if r.get("is_anonymous"))

    metrics = [
        ("Total no Google My Business", data.get("total_review_count", total)),
        ("Avaliações coletadas via API", total),
        ("Nota Média", f"{data.get('average_rating', 0):.2f} ⭐"),
        ("Com texto", with_text),
        ("Com resposta do proprietário", with_reply),
        ("Com fotos/mídia", with_media),
        ("Avaliações anônimas", anon_count),
        ("Taxa de resposta", f"{with_reply/total*100:.1f}%" if total else "0%"),
    ]

    for i, (k, v) in enumerate(metrics, 5):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v
        bg = C_LIGHT if i % 2 == 0 else C_WHITE
        dat(ws[f"A{i}"], bold=True, bg=bg)
        dat(ws[f"B{i}"], center=True, bg=bg)

    # Distribuição de estrelas
    ws.merge_cells("A14:B14")
    ws["A14"].value = "Distribuição de Estrelas"
    hdr(ws["A14"], bg=C_DARK)

    from collections import Counter
    dist = Counter(r["rating"] for r in reviews)
    for i, stars in enumerate(range(5, 0, -1), 15):
        count = dist.get(stars, 0)
        pct = count / total * 100 if total else 0
        ws[f"A{i}"] = "⭐" * stars
        ws[f"B{i}"] = f"{count} ({pct:.1f}%)"
        bg = C_LIGHT if i % 2 == 0 else C_WHITE
        dat(ws[f"A{i}"], center=True, bg=bg)
        dat(ws[f"B{i}"], center=True, bg=bg)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 15

    # ── Aba 2: Avaliações ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Avaliações")

    headers = ["ID", "Autor", "Anônimo", "Nota", "Comentário",
               "Data", "Última Atualização", "Resposta do Proprietário",
               "Data da Resposta", "Tem Mídia"]
    widths  = [20, 25, 10, 8, 70, 22, 22, 60, 22, 12]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws2.cell(row=1, column=col, value=h)
        hdr(cell)
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[1].height = 25
    ws2.freeze_panes = "A2"

    for row_i, r in enumerate(reviews, 2):
        stars_str = "⭐" * r["rating"] if r["rating"] else "—"
        row_data = [
            r.get("review_id", ""),
            r.get("author_name", ""),
            "Sim" if r.get("is_anonymous") else "Não",
            stars_str,
            r.get("comment", ""),
            r.get("create_time", ""),
            r.get("update_time", ""),
            r.get("owner_reply", ""),
            r.get("owner_reply_time", ""),
            "✓" if r.get("has_media") else "",
        ]
        bg = C_LIGHT if row_i % 2 == 0 else C_WHITE
        # Colorir linhas de 5 estrelas em verde claro, 1 estrela em vermelho
        if r["rating"] == 5:
            bg = C_GREEN
        elif r["rating"] == 1:
            bg = C_RED

        for col_i, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_i, column=col_i, value=val)
            dat(cell, bg=bg, center=(col_i in [3, 4, 9, 10]))
        ws2.row_dimensions[row_i].height = 55

    # ── Aba 3: Análise por Nota ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Análise por Nota")

    ws3.merge_cells("A1:D1")
    ws3["A1"].value = "Análise Detalhada por Nota"
    hdr(ws3["A1"], bg=C_DARK, sz=13)
    ws3.row_dimensions[1].height = 30

    h3 = ["Nota", "Qtd", "% do Total", "Com Resposta"]
    for col, h in enumerate(h3, 1):
        cell = ws3.cell(row=2, column=col, value=h)
        hdr(cell, bg=C_GOLD)

    for row_i, stars in enumerate(range(5, 0, -1), 3):
        subset = [r for r in reviews if r["rating"] == stars]
        count = len(subset)
        pct = count / total * 100 if total else 0
        replied = sum(1 for r in subset if r.get("owner_reply"))
        bg = C_LIGHT if row_i % 2 == 0 else C_WHITE

        ws3.cell(row=row_i, column=1, value="⭐" * stars)
        ws3.cell(row=row_i, column=2, value=count)
        ws3.cell(row=row_i, column=3, value=f"{pct:.1f}%")
        ws3.cell(row=row_i, column=4, value=f"{replied} ({replied/count*100:.0f}%)" if count else "—")
        for col in range(1, 5):
            dat(ws3.cell(row=row_i, column=col), bg=bg, center=True)

    for col, w in zip(range(1, 5), [18, 10, 15, 20]):
        ws3.column_dimensions[get_column_letter(col)].width = w

    wb.save(filepath)
    size_kb = Path(filepath).stat().st_size / 1024
    print(f"  ✅ XLSX: {filepath} ({size_kb:.1f} KB)")


# ─── Execução Principal ───────────────────────────────────────────────────────

def main():
    print("\n" + "═" * 70)
    print("  🗺️  Google My Business Reviews API — Embaixada Carioca")
    print(f"  Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print("═" * 70)

    # 1. Autenticar
    print("\n🔐 Autenticando com a Google Business Profile API...")
    creds = get_credentials()
    print("  ✅ Autenticação concluída")

    # 2. Construir cliente da API
    # A Google My Business API v4 usa o endpoint mybusiness.googleapis.com
    # com o discovery document customizado
    print("\n🔧 Conectando à API...")
    try:
        service = googleapiclient.discovery.build(
            "mybusiness",
            "v4",
            credentials=creds,
            discoveryServiceUrl=(
                "https://mybusiness.googleapis.com/$discovery/rest?version=v4"
            ),
            static_discovery=False
        )
        print("  ✅ Conexão estabelecida")
    except Exception as e:
        print(f"  ❌ Erro ao conectar: {e}")
        print("\n  Tentando método alternativo via requests...")
        service = None

    if service is None:
        print("❌ Não foi possível conectar à API. Verifique as credenciais.")
        sys.exit(1)

    # 3. Encontrar a conta e o local
    accounts = get_accounts(service)
    print(f"\n📋 Contas encontradas: {len(accounts)}")
    for acc in accounts:
        print(f"   • {acc.get('accountName', acc.get('name', ''))}")

    account_name, location_name, location_title = find_embaixada_location(
        service, accounts
    )

    # 4. Baixar todas as avaliações
    raw_reviews, avg_rating, total_count = fetch_all_reviews(
        service, account_name, location_name
    )

    # 5. Normalizar os dados
    print("\n🔄 Processando avaliações...")
    parsed = [parse_review(r) for r in raw_reviews]

    # 6. Montar estrutura de saída
    output = {
        "metadata": {
            "location_name": location_title,
            "location_id": location_name,
            "account_id": account_name,
            "extracted_at": datetime.now().isoformat(),
            "total_review_count": total_count,
            "average_rating": round(avg_rating, 2),
            "reviews_collected": len(parsed)
        },
        "average_rating": round(avg_rating, 2),
        "total_review_count": total_count,
        "reviews": parsed
    }

    # 7. Exportar
    print(f"\n📁 Exportando {len(parsed)} avaliações...")
    export_json(output, OUTPUT_JSON)
    export_csv(parsed, OUTPUT_CSV)
    export_xlsx(output, OUTPUT_XLSX)

    # 8. Resumo final
    from collections import Counter
    dist = Counter(r["rating"] for r in parsed)
    with_reply = sum(1 for r in parsed if r.get("owner_reply"))

    print(f"\n{'═'*70}")
    print(f"  ✅ EXTRAÇÃO CONCLUÍDA!")
    print(f"{'═'*70}")
    print(f"  Local:              {location_title}")
    print(f"  Nota média:         {avg_rating:.2f} ⭐")
    print(f"  Total no Google:    {total_count}")
    print(f"  Coletadas via API:  {len(parsed)}")
    print(f"  Com resposta:       {with_reply} ({with_reply/len(parsed)*100:.1f}%)" if parsed else "")
    print(f"\n  Distribuição:")
    for stars in range(5, 0, -1):
        bar = "█" * dist.get(stars, 0)
        print(f"    {'⭐'*stars:<15} {dist.get(stars,0):4d}  {bar[:40]}")
    print(f"\n  Arquivos gerados:")
    print(f"    📄 {OUTPUT_JSON}")
    print(f"    📊 {OUTPUT_CSV}")
    if HAS_OPENPYXL:
        print(f"    📗 {OUTPUT_XLSX}")
    print("═" * 70 + "\n")


if __name__ == "__main__":
    main()

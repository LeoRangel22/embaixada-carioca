#!/usr/bin/env python3
"""
═══════════════════════════════════════════════════════════════════════════════
  Google Maps Reviews Scraper — Embaixada Carioca
  Extrai avaliações via Manus API com Structured Output
  Autor: Manus AI | Data: 2026-06-11
═══════════════════════════════════════════════════════════════════════════════

COMO USAR:
────────────────────────────────────────────────────────────────────────────
1. Obtenha sua chave de API Manus em: https://manus.ai/settings/api
2. Defina a variável de ambiente:
       export MANUS_API_KEY="sua-chave-aqui"
   Ou passe diretamente ao executar:
       MANUS_API_KEY="sua-chave" python3 google_reviews_scraper.py

3. Execute o script:
       python3 google_reviews_scraper.py

4. Aguarde a conclusão (5–15 minutos dependendo do volume de avaliações)
   O script exibe o progresso em tempo real.

5. Os resultados são exportados para:
       reviews_embaixada_carioca.json   (dados completos)
       reviews_embaixada_carioca.csv    (tabela simples)
       reviews_embaixada_carioca.xlsx   (planilha formatada)
────────────────────────────────────────────────────────────────────────────

DEPENDÊNCIAS:
    pip install requests openpyxl
═══════════════════════════════════════════════════════════════════════════════
"""

import os
import sys
import json
import time
import csv
from datetime import datetime

# ─── Verificar dependências ───────────────────────────────────────────────────
try:
    import requests
except ImportError:
    print("❌ 'requests' não instalado. Execute: pip install requests")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("⚠️  'openpyxl' não instalado. O arquivo .xlsx não será gerado.")
    print("   Para instalar: pip install openpyxl")

# ─── Configuração ─────────────────────────────────────────────────────────────

MANUS_API_KEY = os.environ.get("MANUS_API_KEY", "")
MANUS_API_BASE = "https://api.manus.ai"

# URLs do Google Maps da Embaixada Carioca
GOOGLE_MAPS_URL = "https://www.google.com/maps/place/Embaixada+Carioca"
GOOGLE_SEARCH_URL = "https://www.google.com/search?q=Embaixada+Carioca+avalia%C3%A7%C3%B5es"
GOOGLE_SHARE_URL = "https://share.google/rzAyYqJTQnXtJPxOo"

# Arquivos de saída
OUTPUT_JSON = "reviews_embaixada_carioca.json"
OUTPUT_CSV = "reviews_embaixada_carioca.csv"
OUTPUT_XLSX = "reviews_embaixada_carioca.xlsx"

# Polling config
POLL_INTERVAL = 10   # segundos entre cada verificação
MAX_WAIT_TIME = 1800  # 30 minutos de timeout máximo

# ─── Schema de Structured Output ─────────────────────────────────────────────

REVIEWS_SCHEMA = {
    "type": "object",
    "properties": {
        "total_reviews": {
            "type": "integer",
            "description": "Número total de avaliações encontradas no Google Maps"
        },
        "average_rating": {
            "type": "number",
            "description": "Nota média geral (ex: 4.8)"
        },
        "rating_distribution": {
            "type": "object",
            "properties": {
                "five_stars": {"type": "integer", "description": "Quantidade de avaliações com 5 estrelas"},
                "four_stars": {"type": "integer", "description": "Quantidade de avaliações com 4 estrelas"},
                "three_stars": {"type": "integer", "description": "Quantidade de avaliações com 3 estrelas"},
                "two_stars": {"type": "integer", "description": "Quantidade de avaliações com 2 estrelas"},
                "one_star": {"type": "integer", "description": "Quantidade de avaliações com 1 estrela"}
            },
            "required": ["five_stars", "four_stars", "three_stars", "two_stars", "one_star"],
            "additionalProperties": False
        },
        "reviews": {
            "type": "array",
            "description": "Lista de todas as avaliações coletadas",
            "items": {
                "type": "object",
                "properties": {
                    "author_name": {
                        "type": "string",
                        "description": "Nome do autor da avaliação"
                    },
                    "rating": {
                        "type": "integer",
                        "description": "Nota de 1 a 5 estrelas"
                    },
                    "date": {
                        "type": ["string", "null"],
                        "description": "Data da avaliação (ex: 'há 2 semanas', '3 de março de 2025')"
                    },
                    "text": {
                        "type": ["string", "null"],
                        "description": "Texto completo da avaliação"
                    },
                    "language": {
                        "type": ["string", "null"],
                        "description": "Idioma da avaliação (pt, en, es, fr, etc.)"
                    },
                    "is_local_guide": {
                        "type": "boolean",
                        "description": "Se o autor é um Guia Local do Google"
                    },
                    "helpful_count": {
                        "type": ["integer", "null"],
                        "description": "Número de pessoas que acharam útil"
                    },
                    "owner_reply": {
                        "type": ["string", "null"],
                        "description": "Resposta do proprietário à avaliação, se houver"
                    },
                    "owner_reply_date": {
                        "type": ["string", "null"],
                        "description": "Data da resposta do proprietário"
                    }
                },
                "required": [
                    "author_name", "rating", "date", "text",
                    "language", "is_local_guide", "helpful_count",
                    "owner_reply", "owner_reply_date"
                ],
                "additionalProperties": False
            }
        },
        "extraction_notes": {
            "type": ["string", "null"],
            "description": "Observações sobre a extração (ex: limitações, avisos)"
        }
    },
    "required": [
        "total_reviews", "average_rating", "rating_distribution",
        "reviews", "extraction_notes"
    ],
    "additionalProperties": False
}

# ─── Prompt da tarefa ─────────────────────────────────────────────────────────

TASK_PROMPT = f"""Acesse o Google Maps e extraia TODAS as avaliações do restaurante Embaixada Carioca (Morro da Urca, Rio de Janeiro, Brasil).

URLs para acessar:
1. {GOOGLE_SHARE_URL}
2. {GOOGLE_SEARCH_URL}

Instruções detalhadas:
1. Acesse o perfil do Google Maps da Embaixada Carioca
2. Clique em "Ver todas as avaliações" ou na seção de avaliações
3. Role a página para baixo para carregar TODAS as avaliações disponíveis (pode haver centenas)
4. Para cada avaliação, colete:
   - Nome do autor
   - Nota (1-5 estrelas)
   - Data da avaliação
   - Texto completo da avaliação (clique em "Mais" para expandir textos longos)
   - Se é Guia Local do Google
   - Número de "útil" se disponível
   - Resposta do proprietário (se houver)
5. Também colete a nota média geral e a distribuição de estrelas
6. Tente coletar o máximo possível de avaliações — role até o final da lista
7. Inclua avaliações em todos os idiomas (português, inglês, espanhol, etc.)

Seja meticuloso e colete o máximo de avaliações possível antes de finalizar.
"""

# ─── Funções de API ───────────────────────────────────────────────────────────

def get_headers():
    return {
        "x-manus-api-key": MANUS_API_KEY,
        "Content-Type": "application/json"
    }


def create_task():
    """Cria a tarefa de extração de avaliações via Manus API."""
    payload = {
        "message": {
            "content": TASK_PROMPT
        },
        "structured_output_schema": REVIEWS_SCHEMA
    }

    resp = requests.post(
        f"{MANUS_API_BASE}/v2/task.create",
        headers=get_headers(),
        json=payload,
        timeout=30
    )
    resp.raise_for_status()
    data = resp.json()

    if not data.get("ok"):
        raise RuntimeError(f"Erro ao criar tarefa: {data.get('error', {})}")

    return data["task_id"]


def poll_task(task_id):
    """
    Faz polling da tarefa até ela completar.
    Retorna o resultado estruturado quando disponível.
    """
    start_time = time.time()
    last_cursor = None
    structured_result = None
    status = "running"
    last_status_print = ""

    print(f"\n⏳ Aguardando o agente coletar as avaliações...")
    print(f"   (Isso pode levar de 5 a 20 minutos dependendo do volume)\n")

    while True:
        elapsed = time.time() - start_time
        if elapsed > MAX_WAIT_TIME:
            raise TimeoutError(f"Timeout após {MAX_WAIT_TIME/60:.0f} minutos")

        # Montar URL de polling
        params = {
            "task_id": task_id,
            "order": "asc",
            "limit": 50
        }
        if last_cursor:
            params["cursor"] = last_cursor

        resp = requests.get(
            f"{MANUS_API_BASE}/v2/task.listMessages",
            headers=get_headers(),
            params=params,
            timeout=30
        )
        resp.raise_for_status()
        data = resp.json()

        if not data.get("ok"):
            raise RuntimeError(f"Erro ao listar mensagens: {data.get('error', {})}")

        messages = data.get("messages", [])

        for msg in messages:
            msg_type = msg.get("type", "")

            # Atualizar cursor para próxima página
            if msg.get("id"):
                last_cursor = msg["id"]

            # Verificar status
            if msg_type == "status_update":
                su = msg.get("status_update", {})
                new_status = su.get("agent_status", status)
                if new_status != last_status_print:
                    status_icons = {
                        "running": "🔄",
                        "stopped": "✅",
                        "waiting": "⏸️",
                        "error": "❌"
                    }
                    icon = status_icons.get(new_status, "❓")
                    elapsed_min = elapsed / 60
                    print(f"  {icon} Status: {new_status.upper()} | Tempo: {elapsed_min:.1f}min")
                    last_status_print = new_status
                status = new_status

            # Capturar resultado estruturado
            elif msg_type == "structured_output_result":
                sor = msg.get("structured_output_result", {})
                if sor.get("success"):
                    structured_result = sor.get("value")
                    print(f"\n  ✅ Structured Output recebido com sucesso!")
                else:
                    print(f"\n  ⚠️  Structured Output com erro: {sor.get('error')}")
                    structured_result = sor.get("value")  # Usar fallback

            # Exibir mensagens do agente
            elif msg_type == "assistant_message":
                content = msg.get("content", "")
                if content and len(content) > 10:
                    preview = content[:120].replace("\n", " ")
                    print(f"  💬 Agente: {preview}...")

        # Verificar se terminou
        if status == "stopped":
            if structured_result is not None:
                return structured_result
            else:
                # Tarefa terminou mas sem structured output ainda — aguardar mais um ciclo
                time.sleep(5)
                # Tentar mais uma vez
                continue

        elif status == "error":
            raise RuntimeError("A tarefa falhou com erro. Verifique o painel do Manus.")

        elif status == "waiting":
            print(f"  ⏸️  Agente aguardando interação — verificando em {POLL_INTERVAL}s...")

        # Aguardar antes do próximo poll
        time.sleep(POLL_INTERVAL)

    return structured_result


# ─── Funções de Exportação ────────────────────────────────────────────────────

def export_json(data, filepath):
    """Exporta os dados para JSON formatado."""
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  ✅ JSON salvo: {filepath}")


def export_csv(reviews, filepath):
    """Exporta as avaliações para CSV."""
    if not reviews:
        print(f"  ⚠️  Nenhuma avaliação para exportar em CSV")
        return

    fieldnames = [
        "autor", "nota", "data", "idioma", "guia_local",
        "util_count", "texto", "resposta_proprietario", "data_resposta"
    ]

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in reviews:
            writer.writerow({
                "autor": r.get("author_name", ""),
                "nota": r.get("rating", ""),
                "data": r.get("date", ""),
                "idioma": r.get("language", ""),
                "guia_local": "Sim" if r.get("is_local_guide") else "Não",
                "util_count": r.get("helpful_count", ""),
                "texto": r.get("text", ""),
                "resposta_proprietario": r.get("owner_reply", ""),
                "data_resposta": r.get("owner_reply_date", "")
            })
    print(f"  ✅ CSV salvo: {filepath}")


def export_xlsx(data, filepath):
    """Exporta as avaliações para Excel formatado profissionalmente."""
    if not HAS_OPENPYXL:
        return

    reviews = data.get("reviews", [])
    if not reviews:
        print(f"  ⚠️  Nenhuma avaliação para exportar em XLSX")
        return

    wb = openpyxl.Workbook()

    # ─── Aba 1: Resumo ────────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Resumo"

    # Cores da identidade visual
    COLOR_HEADER = "00405A"   # Azul escuro
    COLOR_ACCENT = "F59B1E"   # Dourado
    COLOR_LIGHT = "E8F4F8"    # Azul claro
    COLOR_WHITE = "FFFFFF"

    def header_style(cell, bg=COLOR_HEADER, fg=COLOR_WHITE, bold=True, size=11):
        cell.font = Font(bold=bold, color=fg, size=size, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def data_style(cell, bold=False, center=False, bg=None):
        cell.font = Font(bold=bold, size=10, name="Calibri")
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        if center:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Título
    ws_summary.merge_cells("A1:D1")
    title_cell = ws_summary["A1"]
    title_cell.value = "Embaixada Carioca — Relatório de Avaliações Google Maps"
    header_style(title_cell, bg=COLOR_HEADER, size=14)
    ws_summary.row_dimensions[1].height = 35

    # Subtítulo
    ws_summary.merge_cells("A2:D2")
    sub_cell = ws_summary["A2"]
    sub_cell.value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    data_style(sub_cell, bg=COLOR_LIGHT, center=True)
    ws_summary.row_dimensions[2].height = 20

    # Métricas principais
    ws_summary["A4"] = "Métrica"
    ws_summary["B4"] = "Valor"
    header_style(ws_summary["A4"], bg=COLOR_ACCENT, fg=COLOR_WHITE)
    header_style(ws_summary["B4"], bg=COLOR_ACCENT, fg=COLOR_WHITE)

    metrics = [
        ("Total de Avaliações", data.get("total_reviews", len(reviews))),
        ("Nota Média", f"{data.get('average_rating', 0):.1f} ⭐"),
        ("Avaliações Coletadas", len(reviews)),
        ("Com Resposta do Proprietário", sum(1 for r in reviews if r.get("owner_reply"))),
        ("Guias Locais", sum(1 for r in reviews if r.get("is_local_guide"))),
        ("Em Português", sum(1 for r in reviews if r.get("language") == "pt")),
        ("Em Inglês", sum(1 for r in reviews if r.get("language") == "en")),
        ("Em Espanhol", sum(1 for r in reviews if r.get("language") == "es")),
    ]

    for i, (metric, value) in enumerate(metrics, start=5):
        ws_summary[f"A{i}"] = metric
        ws_summary[f"B{i}"] = value
        bg = COLOR_LIGHT if i % 2 == 0 else COLOR_WHITE
        data_style(ws_summary[f"A{i}"], bold=True, bg=bg)
        data_style(ws_summary[f"B{i}"], center=True, bg=bg)

    # Distribuição de estrelas
    dist = data.get("rating_distribution", {})
    ws_summary["A14"] = "Distribuição de Estrelas"
    ws_summary.merge_cells("A14:B14")
    header_style(ws_summary["A14"], bg=COLOR_HEADER)

    star_data = [
        ("⭐⭐⭐⭐⭐ (5 estrelas)", dist.get("five_stars", 0)),
        ("⭐⭐⭐⭐ (4 estrelas)", dist.get("four_stars", 0)),
        ("⭐⭐⭐ (3 estrelas)", dist.get("three_stars", 0)),
        ("⭐⭐ (2 estrelas)", dist.get("two_stars", 0)),
        ("⭐ (1 estrela)", dist.get("one_star", 0)),
    ]

    for i, (label, count) in enumerate(star_data, start=15):
        ws_summary[f"A{i}"] = label
        ws_summary[f"B{i}"] = count
        bg = COLOR_LIGHT if i % 2 == 0 else COLOR_WHITE
        data_style(ws_summary[f"A{i}"], bg=bg)
        data_style(ws_summary[f"B{i}"], center=True, bg=bg)

    ws_summary.column_dimensions["A"].width = 35
    ws_summary.column_dimensions["B"].width = 20

    # ─── Aba 2: Avaliações ────────────────────────────────────────────────────
    ws_reviews = wb.create_sheet("Avaliações")

    headers = [
        "Autor", "Nota", "Data", "Idioma", "Guia Local",
        "Útil", "Texto da Avaliação", "Resposta do Proprietário", "Data da Resposta"
    ]
    col_widths = [25, 8, 20, 10, 12, 8, 60, 50, 20]

    # Cabeçalho
    for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws_reviews.cell(row=1, column=col, value=header)
        header_style(cell, bg=COLOR_HEADER)
        ws_reviews.column_dimensions[get_column_letter(col)].width = width
    ws_reviews.row_dimensions[1].height = 25

    # Dados
    for row_idx, review in enumerate(reviews, start=2):
        stars = "⭐" * review.get("rating", 0)
        row_data = [
            review.get("author_name", ""),
            stars,
            review.get("date", ""),
            review.get("language", "").upper() if review.get("language") else "",
            "✓ Guia Local" if review.get("is_local_guide") else "",
            review.get("helpful_count", ""),
            review.get("text", ""),
            review.get("owner_reply", ""),
            review.get("owner_reply_date", "")
        ]

        bg = COLOR_LIGHT if row_idx % 2 == 0 else COLOR_WHITE
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_reviews.cell(row=row_idx, column=col_idx, value=value)
            data_style(cell, bg=bg, center=(col_idx in [2, 3, 4, 5, 6]))

        ws_reviews.row_dimensions[row_idx].height = 60

    # Freeze header
    ws_reviews.freeze_panes = "A2"

    # ─── Aba 3: Análise por Idioma ────────────────────────────────────────────
    ws_lang = wb.create_sheet("Por Idioma")

    from collections import Counter
    lang_counts = Counter(r.get("language", "desconhecido") or "desconhecido" for r in reviews)
    lang_ratings = {}
    for r in reviews:
        lang = r.get("language", "desconhecido") or "desconhecido"
        if lang not in lang_ratings:
            lang_ratings[lang] = []
        lang_ratings[lang].append(r.get("rating", 0))

    ws_lang["A1"] = "Análise por Idioma"
    ws_lang.merge_cells("A1:D1")
    header_style(ws_lang["A1"], bg=COLOR_HEADER, size=13)
    ws_lang.row_dimensions[1].height = 30

    lang_headers = ["Idioma", "Avaliações", "Nota Média", "% do Total"]
    for col, header in enumerate(lang_headers, start=1):
        cell = ws_lang.cell(row=2, column=col, value=header)
        header_style(cell, bg=COLOR_ACCENT)

    total = len(reviews)
    for row_idx, (lang, count) in enumerate(sorted(lang_counts.items(), key=lambda x: -x[1]), start=3):
        avg = sum(lang_ratings[lang]) / len(lang_ratings[lang]) if lang_ratings[lang] else 0
        pct = (count / total * 100) if total > 0 else 0
        bg = COLOR_LIGHT if row_idx % 2 == 0 else COLOR_WHITE

        ws_lang.cell(row=row_idx, column=1, value=lang.upper()).fill = PatternFill("solid", fgColor=bg)
        ws_lang.cell(row=row_idx, column=2, value=count).fill = PatternFill("solid", fgColor=bg)
        ws_lang.cell(row=row_idx, column=3, value=f"{avg:.1f} ⭐").fill = PatternFill("solid", fgColor=bg)
        ws_lang.cell(row=row_idx, column=4, value=f"{pct:.1f}%").fill = PatternFill("solid", fgColor=bg)

    for col in range(1, 5):
        ws_lang.column_dimensions[get_column_letter(col)].width = 20

    wb.save(filepath)
    print(f"  ✅ Excel salvo: {filepath}")


# ─── Execução Principal ───────────────────────────────────────────────────────

def main():
    print("\n" + "═"*70)
    print("  🗺️  Google Maps Reviews Scraper — Embaixada Carioca")
    print(f"  Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print("═"*70)

    # Verificar API Key
    if not MANUS_API_KEY:
        print("\n❌ MANUS_API_KEY não definida!")
        print("\nPara obter sua chave de API:")
        print("  1. Acesse: https://manus.ai/settings/api")
        print("  2. Crie uma nova API Key")
        print("  3. Execute: export MANUS_API_KEY='sua-chave-aqui'")
        print("  4. Execute o script novamente")
        sys.exit(1)

    print(f"\n🔑 API Key: {MANUS_API_KEY[:8]}...{MANUS_API_KEY[-4:]}")

    # Criar tarefa
    print("\n📋 Criando tarefa de extração de avaliações...")
    try:
        task_id = create_task()
        print(f"  ✅ Tarefa criada: {task_id}")
        print(f"  🔗 Acompanhe em: https://manus.ai/app/tasks/{task_id}")
    except Exception as e:
        print(f"  ❌ Erro ao criar tarefa: {e}")
        sys.exit(1)

    # Aguardar e coletar resultado
    try:
        result = poll_task(task_id)
    except TimeoutError as e:
        print(f"\n⏰ {e}")
        print("Tente aumentar MAX_WAIT_TIME no script ou execute novamente.")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ Erro durante polling: {e}")
        sys.exit(1)

    if not result:
        print("\n❌ Nenhum resultado obtido.")
        sys.exit(1)

    # Exibir resumo
    reviews = result.get("reviews", [])
    print(f"\n{'═'*70}")
    print(f"  📊 RESULTADO DA EXTRAÇÃO")
    print(f"{'═'*70}")
    print(f"  Total de avaliações coletadas: {len(reviews)}")
    print(f"  Nota média declarada: {result.get('average_rating', 'N/A')}")
    print(f"  Total no Google Maps: {result.get('total_reviews', 'N/A')}")
    if result.get("extraction_notes"):
        print(f"  Notas: {result['extraction_notes']}")

    # Exportar resultados
    print(f"\n📁 Exportando resultados...")
    export_json(result, OUTPUT_JSON)
    export_csv(reviews, OUTPUT_CSV)
    export_xlsx(result, OUTPUT_XLSX)

    print(f"\n{'═'*70}")
    print(f"  ✅ Extração concluída!")
    print(f"  Arquivos gerados:")
    print(f"    📄 {OUTPUT_JSON}")
    print(f"    📊 {OUTPUT_CSV}")
    if HAS_OPENPYXL:
        print(f"    📗 {OUTPUT_XLSX}")
    print("═"*70 + "\n")


if __name__ == "__main__":
    main()
